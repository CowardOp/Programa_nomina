from datetime import datetime, time
import holidays
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from calendar import monthrange


class procesador_excel:
    def __init__(self):
        self.color_index_to_rgb = {
            3: "FF0000",
            33: "00B0F0",
            47: "7030A0",
            43: "92D050",
            6: "FFFF00",
            49: "002060",
            14: "00B050",
            16: "808080",
            44: "FFC000",
        }

        self.col_destino_color = {
            3: 104,
            33: 105,
            47: 106,
            43: 107,
            6: 108,
            49: 109,
            14: 110,
            16: 111,
            44: 112,
        }

    def contar_colores(self, ws, fila):
        colores = {idx: 0 for idx in self.color_index_to_rgb}
        for col in range(4, 66):
            celda = ws.cell(row=fila, column=col)
            fill = celda.fill
            if fill and fill.start_color and fill.start_color.type == "rgb":
                color = fill.start_color.rgb
                if isinstance(color, str):
                    color = color[-6:].upper()
                    for idx, rgb in self.color_index_to_rgb.items():
                        if color == rgb:
                            colores[idx] += 1
        return colores

    def calcular_horas_y_colores(self, ruta_archivo, quincena, mes, anio):
        import os

        wb = load_workbook(ruta_archivo, keep_vba=True, data_only=True)
        ws = wb.worksheets[0]  # Hoja1

        valor_hora_extra = ws.cell(row=10, column=189).value or 0

        for fila in range(7, ws.max_row + 1):
            total_horas = 0
            total_extras = 0

            for idx, col in enumerate(range(4, 66, 4)):
                entrada = ws.cell(row=fila, column=col).value
                salida = ws.cell(row=fila, column=col + 1).value

                if isinstance(entrada, (datetime, time)) and isinstance(
                    salida, (datetime, time)
                ):
                    hora_entrada = entrada.hour + entrada.minute / 60
                    hora_salida = salida.hour + salida.minute / 60
                    horas_trabajadas = hora_salida - hora_entrada

                    if 6 <= hora_entrada < 12:
                        horas_trabajadas -= 1.5
                    elif 12 <= hora_entrada < 23:
                        horas_trabajadas -= 0.25

                    if horas_trabajadas < 0:
                        horas_trabajadas = 0

                    col_horas = 68 + idx
                    ws.cell(row=fila, column=col_horas).value = round(
                        horas_trabajadas / 24, 6
                    )
                    total_horas += horas_trabajadas / 24

                    if horas_trabajadas > 8:
                        total_extras += horas_trabajadas - 8

            ws.cell(row=fila, column=84).value = round(total_horas, 6)
            # Calcular penalización por días con menos de 7.33 horas trabajadas
            penalizacion_faltante = 0
            for col in range(68, 84):  # Columnas que contienen horas diarias
                horas_dia = ws.cell(row=fila, column=col).value
                if isinstance(horas_dia, (int, float)):
                    horas_dia_real = horas_dia * 24  # convertir a horas reales
                    if horas_dia_real < 8:
                        penalizacion_faltante += 8 - horas_dia_real

            # Marcar con 1 en columnas 86 a 101 si hay datos en 68 a 83
            # Calcular diferencia con 7.33 y colocar en columnas 86 a 101
            for idx, col_base in enumerate(range(68, 84)):
                valor = ws.cell(row=fila, column=col_base).value
                col_diferencia = 86 + idx  # columnas 86 a 101

                if isinstance(valor, (int, float)):
                    horas_reales = valor * 24  # convertir a horas
                    diferencia = round(horas_reales - 8, 2)
                    ws.cell(row=fila, column=col_diferencia).value = diferencia
                else:
                    ws.cell(row=fila, column=col_diferencia).value = ""
                    # Nueva lógica para columna 102: sumar las diferencias de columnas 86 a 101
            suma_diferencias = 0
            for col in range(86, 102):  # columnas 86 a 101
                valor = ws.cell(row=fila, column=col).value
                if isinstance(valor, (int, float)):
                    suma_diferencias += valor

            ws.cell(row=fila, column=102).value = round(suma_diferencias, 2)
            ws.cell(row=fila, column=115).value = ws.cell(row=fila, column=102).value
            # Calcular valor en pesos de las horas extras (columna 115) y ponerlo en 116
            salario_minimo = 1423500
            valor_hora_ordinaria = salario_minimo / 220
            valor_hora_extra_diurna = valor_hora_ordinaria * 1.25

            horas_extras = ws.cell(row=fila, column=115).value

            if isinstance(horas_extras, (int, float)):
                valor_total_extras = horas_extras * valor_hora_extra_diurna
                if valor_total_extras > 0:
                    ws.cell(row=fila, column=116).value = round(valor_total_extras, 2)
                else:
                    ws.cell(row=fila, column=116).value = ""
            else:
                ws.cell(row=fila, column=116).value = ""

            dias_trabajados = 0
            for c in range(68, 84):
                valor = ws.cell(row=fila, column=c).value
                if isinstance(valor, (int, float)) and valor > 0:
                    dias_trabajados += 1
            ws.cell(row=fila, column=103).value = dias_trabajados

            ws.cell(row=fila, column=113).value = (
                f"=CY{fila}+CZ{fila}+DA{fila}+DB{fila}+DC{fila}+DD{fila}+DE{fila}+DF{fila}+DG{fila}+DH{fila}"
            )

            ws.cell(row=fila, column=114).value = 711750
            valor_total_extras = total_extras * valor_hora_extra
            # ws.cell(row=fila, column=116).value = round(valor_total_extras, 2)

            festivos = sum(
                1 for col in range(4, 66) if ws.cell(row=fila, column=col).value == "F"
            )
            dominicales = sum(
                1 for col in range(4, 66) if ws.cell(row=fila, column=col).value == "D"
            )
            ws.cell(row=fila, column=117).value = festivos
            ws.cell(row=fila, column=118).value = dominicales

            contadores = self.contar_colores(ws, fila)
            for idx, cantidad in contadores.items():
                col = self.col_destino_color[idx]
                ws.cell(row=fila, column=col).value = (
                    cantidad if cantidad != 0 else None
                )

            ws.cell(row=fila, column=122).value = (
                f"=DJ{fila}+DL{fila}+DN{fila}+DO{fila}"
            )
            ws.cell(row=fila, column=124).value = f"=DR{fila}-DS{fila}"

        # === Pintar CH a CV (cols 87–99) según la quincena ===
        AZUL = PatternFill(
            start_color="00B0F0", end_color="00B0F0", fill_type="solid"
        )  # Día normal
        NARANJA = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )  # Domingo

        meses = {
            "Enero": 1,
            "Febrero": 2,
            "Marzo": 3,
            "Abril": 4,
            "Mayo": 5,
            "Junio": 6,
            "Julio": 7,
            "Agosto": 8,
            "Septiembre": 9,
            "Octubre": 10,
            "Noviembre": 11,
            "Diciembre": 12,
        }
        mes_num = meses.get(mes, 1)

        if quincena == "Primera":
            dia_inicio, dia_fin = 1, 15
        else:
            dia_inicio, dia_fin = 16, monthrange(anio, mes_num)[1]

        for fila in range(7, ws.max_row + 1):
            col_inicio = 86
            dia_actual = dia_inicio
            colombian_holidays = holidays.Colombia(years=anio)

        for fila in range(7, ws.max_row + 1):
            col_inicio = 86
            dia_actual = dia_inicio
            for col in range(col_inicio, col_inicio + (dia_fin - dia_inicio + 1)):
                try:
                    fecha = datetime(anio, mes_num, dia_actual).date()
                    if fecha in colombian_holidays or fecha.weekday() == 6:
                        ws.cell(row=fila, column=col).fill = NARANJA
                    else:
                        ws.cell(row=fila, column=col).fill = AZUL
                    dia_actual += 1
                except ValueError:
                    continue
                    # === Sumar los "1" de columnas 86 a 101 que tengan fondo naranja ===
            total_domingos_trabajados = 0
            for col in range(86, 102):
                celda = ws.cell(row=fila, column=col)
                fill = celda.fill
                color = fill.start_color.rgb if fill and fill.start_color else None
                if color and color[-6:].upper() == "FFA500" and celda.value != None:
                    total_domingos_trabajados += 1
            ws.cell(row=fila, column=117).value = total_domingos_trabajados

        # Guardar
        base, ext = os.path.splitext(ruta_archivo)
        nuevo_nombre = f"{base}_Procesado{ext}"
        wb.save(nuevo_nombre)
        print(f"Archivo guardado como: {nuevo_nombre}")
